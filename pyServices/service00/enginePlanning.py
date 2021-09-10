from openpyxl import load_workbook
from openpyxl import Workbook
import pymysql


import datetime


IP = "192.168.1.191"
PASS = "tk2718"


#db = pymysql.connect(IP, "brch", PASS, "mes")
db = pymysql.connect(host="IP", user="brch", password="tk2718", db="mes")


class LineHolder:
    def __init__(self, id_line, line, id_area):
        self.id = id_line
        self.line = line
        self.id_area = id_area

    def printObject(self):
        print("ID: %d\tName: %s\tArea:%d" % (self.id, self.line, self.id_area))


class PresentationHolder:
    def __init__(self, code, id_presentation, descripcion, undMin):
        self.code = code
        self.id_presentation = id_presentation
        self.descripcion = descripcion
        self.undMin = undMin

    def printObject(self):
        print("Code: %s\tId[%s] Description: %s\tund/min%s" %
              (self.code, self.id_presentation, self.descripcion, self.undMin))


class Plan:
    def __init__(self, planned, produced, turn, nominal_speed, id_presentation, id_line, date_planning, version):
        self.planned = planned
        self.produced = produced
        self.turn = turn
        self.nominal_speed = nominal_speed
        self.id_presentation = id_presentation
        self.id_line = id_line
        self.date_planning = date_planning
        self.version = version

    def printObject(self):
        print("planned: %d\tproduced: %d\tturn: %d\tnominalSpeed: %f\tpresentation: %d\tline: %d\tdatePlanning: %s\tversion: %d" %
              (self.planned, self.produced, self.turn, self.nominal_speed, self.id_presentation, self.id_line, self.date_planning, self.version))


class PlanHolder:

    def __init__(self, line, code, presentation, day, turn, planned, produced):
        self.line = line
        self.code = code
        self.presentation = presentation
        self.day = day
        self.turn = turn
        self.planned = planned
        self.produced = produced

    def printObject(self):
        print(
            "ID[%d] Line: %s\tCode[%d]\t ID[%d] Presentation: %s\t\tDay: %s\tT: %d\tPlanned: %d\tProduced: %d\tnominalSpeed: %f" %
            (self.line_id, self.line, self.code, self.presentation_id, self.presentation, self.day, self.turn, self.planned, self.produced, self.nominal_speed))

    def setLineID(self, pline_id):
        self.line_id = pline_id

    def setPresentationID(self, ppresentation_id):
        self.presentation_id = ppresentation_id

    def setNominalSpeed(self, nominal_speed):
        self.nominal_speed = nominal_speed


def cleanWokbook(file_template, dateWeek):

    # daysOfWeek = ["L", "K", "M", "J", "V", "S", "D"]
    daysOfWeek = []
    prePlanning = []

    startDay = dateWeek  # Moday of week 19

    for i in range(7):
        calc = startDay + datetime.timedelta(days=i)
        stringDate = '{:%Y-%m-%d}'.format(calc)

        daysOfWeek.append(stringDate)

    workbook = load_workbook(filename=file_template)
    sheet = workbook.worksheets[0]
    current = 0
    count = 0
    current_row = 1
    Lines = []

    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=26, values_only=True):
        l = row[0]

        if type(l) is str:
            if not l in Lines:
                cleanLine = l.split("(")[0]
                Lines.append(cleanLine)

                # print("--", Lines[current], "---")
                # print("\n")
                current += 1
        else:

            presentation = row[1]
            code = row[0]
            j = 0
            day = 0
            currentTurn = 0
            prod = ""

            line = Lines[current-1]

            for i in range(4, len(row)):
                planned = row[i]
                next_row = current_row+current+1
                next_col = i + 1
                produced = sheet.cell(row=(next_row), column=next_col).value if type(
                    sheet.cell(row=(next_row), column=next_col).value) is int else 0

                if type(row[i]) is int and "VERS" not in row[1]:

                    prePlanning.append(PlanHolder(
                        line, code, presentation, daysOfWeek[day-1], currentTurn, planned, produced))

                    count += 1

                if j % 3 == 0:
                    day += 1
                    currentTurn = 0

                currentTurn += 1
                j += 1

            current_row += 1

    workbook.close()

    print("\t\t--CleanDone--")

    return prePlanning


def getLines(file_template):
    LinesHolder = []

    workbook = load_workbook(filename=file_template)
    sheet = workbook.worksheets[2]

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        LinesHolder.append(LineHolder(row[0], row[1], row[2]))

    workbook.close()

    return LinesHolder


def getPresentations(file_template):
    
    PresentationCatalog = []
    workbook = load_workbook(filename=file_template)
    sheet = workbook.worksheets[1]

    count = 0

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
            PresentationCatalog.append(
                PresentationHolder(row[0], row[1], row[2], row[4]))

    workbook.close()

    return PresentationCatalog


def getLineIdByName(L, name):
    line_id = -1

    for l in L:
        if l.line == name:
            line_id = l.id

    return line_id


def getPresentationByCode(L, code):
    presentation = PresentationHolder(code, -1, "", 1)

    # print("-"+code+"-\n")

    for p in L:
        if p.code == code:
            presentation = p

    return presentation



def insertPlan(pline, ppresentation, pdate_planning, pversion, pturn, pnominal_speed, pproduced, pplanned):
    db = pymysql.connect(host="IP", user="brch", password="tk2718", db="mes")

    line = str(pline)
    presentation = str(ppresentation) 
    date_planning = str(pdate_planning).replace("-", "")
    version = str(pversion)
    turn   = str(pturn)
    nominal_speed = str(pnominal_speed)
    produced = str(pproduced)
    planned = str(pplanned)


    # print("line: "+line+" presetation: "+presentation+" date: "+date_planning+" version: "+version+" turn: "+turn+" nominal_speed: "+nominal_speed+" produced: "+produced+" planned: "+planned)

    cursor = db.cursor()
    sql = "INSERT INTO Planning(id_line, id_presentation, date_planning, version, turn, nominal_speed, produced, planned)"\
          " VALUES(%s, %s, %s, %s, %s, %s, %s, %s)"

    cursor.execute(sql, (line, presentation, date_planning, version, turn, nominal_speed, produced, planned))

    db.commit()
    cursor.close()

    db.close()

    return "Cargado!"


def deletePlanRange(begin, end):
    db = pymysql.connect(host="IP", user="brch", password="tk2718", db="mes")

    format = "%Y%m%d"

    b = begin.strftime(format)
    e = end.strftime(format)

    cursor = db.cursor()
    sql = "delete from Planning where date_planning between "+ b +" and "+e

    cursor.execute(sql)

    cursor.close()
    db.close()


def beginDate(year, week):


    first = datetime.datetime(year, 1, 1)

    first = first - datetime.timedelta(days=first.weekday())
    
    begin = first + datetime.timedelta(days=7*(week-1))

    return begin

def endDate(begin):

    end = begin + datetime.timedelta(days=6)

    return end

def initPlanningEngine(file_template, dateWeek):
    plan = []
    data = []
    notInDB = []

    endWeek = endDate(dateWeek)


    print("begin: "+dateWeek.isoformat()+" end: "+endWeek.isoformat())


    print(deletePlanRange(dateWeek, endWeek))

    preProcess = cleanWokbook(file_template, dateWeek)

    linesCatalog = getLines(file_template)
    presentationCatalog = getPresentations(file_template)


    # Match presentation and lines propeties
    for p in preProcess:
        presentation = getPresentationByCode(
            presentationCatalog, p.code)

        p.setLineID(getLineIdByName(linesCatalog, p.line))
        p.setPresentationID(presentation.id_presentation)
        p.setNominalSpeed(presentation.undMin)

        

        if presentation.id_presentation == -1:
            
            if presentation.code not in notInDB:
                notInDB.append(presentation.code)
                print("Presentation not in databse:\t%s" % (presentation.code))
            

    # set FInal objet
    for p in preProcess:
        if p.presentation_id != -1:
            plan.append(Plan(p.planned, p.produced, p.turn,
                         p.nominal_speed, p.presentation_id, p.line_id, p.day, 1))
    for d in plan:
        data.append((d.planned, d.produced, d.turn, d.nominal_speed, d.id_presentation,
                     d.id_line, d.date_planning, d.version))

        insertPlan(d.id_line, d.id_presentation, d.date_planning, d.version, d.turn, d.nominal_speed, d.produced, d.planned)

    # planResult = Workbook()  # set file path

    # hoja = planResult.active

    # # hoja.append(('planned', 'produced', 'turn', 'nominal_speed', 'id_presentation',
    # #              'id_line', 'date_planning', 'version'))

    # for row in data:
    #     # data.printObject()
    #     hoja.append(row)  # save file

    # stamp = datetime.datetime.now().isoformat()

    # stamp = stamp.replace(":", "")
    # stamp = stamp.replace(".", "")
    
    # planResult.save("resultFile/planning"+stamp+".xlsx")
    # planResult.close()
    print("Done!")
    return notInDB


# if __name__ == "__main__":
#     pass

#     begin = beginDate(2020, 22)
#     fileName = "templatePlanning.xlsx"

#     initPlanningEngine(fileName, begin)
    