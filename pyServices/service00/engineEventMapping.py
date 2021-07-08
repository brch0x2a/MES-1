import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

import time
import math

IP = "192.168.1.191"
PASS = "tk2718"

FILE_NAME = "static/docs/ValidacionOEE.xlsx"

class EventXLineHolder:
    def __init__(self, id_transaction, date_event, id_line, line, id_event, sub_category, branch, event, minutes, note):


        self.id_transaction = id_transaction
        self.date_event = date_event
        self.id_line = id_line
        self.line = line
        self.id_event = id_event
        self.sub_category = sub_category
        self.branch = branch
        self.event = event
        self.minutes = minutes
        self.note = note


    def printObject(self):
        print("[%d] %s\tline[%d] %s\tevent[%d] %s|%s|%s\tminutes:%d" %
         (self.id_transaction, self.date_event, self.id_line, self.line, self.id_event, self.sub_category, self.branch, self.event, self.minutes))



class Line:

    def __init__(self, id, name):
        self.id = id
        self.name = name

    def printObject(self):
        print("[%d] %s" % (self.id, self.name))


class EventHolder:


    def __init__(self, code, day, turn, minute):
        self.code = code 
        self.day = day 
        self.turn = turn 
        self.minute = minute

    def setTransaction(self, transaction):
        self.transaction = transaction


    def print(self):
        print("(%d, %d, %d, %f)"%(self.code, self.day, self.turn, self.minute))


class EventMapper:

    def __init__(self, line):
        self.line = line
        self.Event = []

    def add(self, eventHolder):

            if  self.isIn(eventHolder) == -1:
                self.Event.append(eventHolder)
            else:
                i = self.isIn(eventHolder)    
                self.Event[i].minute += eventHolder.minute

    def isIn(self, eventHolder):
        index = -1

        for i in range(len(self.Event)):
            if eventHolder.code == self.Event[i].code and eventHolder.day == self.Event[i].day and eventHolder.turn == self.Event[i].turn:
                index = i
            
        return index

    def print(self):
        print("Line: %s"%(self.line))
        print("Code | Day | Turn | Minute")
        for e in self.Event:
            print("(%d, %d, %d, %d)"%(e.code, e.day, e.turn, e.minute))



class LineMapper:

    def __init__(self, name):

        self.name = name
        self.sheet = []
        self.codes = [] 

    def add(self, exl):

        if exl.id_event not in self.codes:
            row = [] 
            row.append(exl)
            self.codes.append(exl.id_event)
            self.sheet.append(row)

        else:
            rowIndex = self.findRow(exl.id_event)

            self.sheet[rowIndex].append(exl)

    
    def findRow(self, code):

        for i in range(len(self.codes)):
            if code == self.codes[i]:
                return i

 
    def print(self):
        for i in range(len(self.sheet)):
            print("[%d]\t"%(self.codes[i]), end="")
            for j in range(len(self.sheet[i])):
                print(" < %d | %s > "%
                (self.sheet[i][j].minutes, self.sheet[i][j].date_event), end="")

            print("\n")


class CodeHolder:
    def __init__(self, row, col, code):
        self.row = row 
        self.col = col
        self.code = code

class DataHolder:

    def __init__(self, row, col, minutes):
        self.row = row 
        self.col = col
        self.minutes = minutes

    def print(self):
        print("[%d][%d]-->%d"%(self.row, self.col, self.minutes))


class RawMapper:

    def __init__(self, line_name):
        self.line_name = line_name
        self.Data = []
        self.codes = [] 

    def isRegister(self, dataHolder):
        result = -1

        for i in range(len(self.Data)):
            if self.Data[i].row == dataHolder.row and self.Data[i].col == dataHolder.col:
                result = i

        return result

    def findCode(self, row):
        for i in range(len(self.codes)):
            if self.codes[i].row == row:
                return i

    def add(self, dataHolder):

        index = self.isRegister(dataHolder) 


        if index != -1:
            # dataHolder.print()
            # print("sum[%d][%d] = %d + %d\n\n"%
            # (dataHolder.row, dataHolder.col, self.Data[index].minutes, dataHolder.minutes))

            self.Data[index].minutes += dataHolder.minutes

        else:
            # print("_________________")
            # dataHolder.print()
            # print("_________________\n")
            self.Data.append(dataHolder)






def findLineLike(name):
    db = pymysql.connect(IP, "brch", PASS, "mes")

    cursor = db.cursor()


    sql = "SELECT id, name FROM mes.Line where name like %s;"

    cursor.execute(sql, (name))

    cursor.close()

    line = cursor.fetchall()

    db.close()

    return line


def getLines():
    db = pymysql.connect(IP, "brch", PASS, "mes")


    cursor = db.cursor()


    sql =   '''
        SELECT 
            *
        FROM
            Line
        WHERE
            id IN (1 , 2,
                3,
                4,
                5,
                6,
                7,
                8,
                9,
                10,
                11,
                12,
                13,
                14,
                15,
                16,
                17,
                18,
                19,
                20)
    '''

    cursor.execute(sql)

    cursor.close()

    lines = cursor.fetchall()

    db.close()

    return lines



def getEvents(begin, end):
    db = pymysql.connect(IP, "brch", PASS, "mes")


    cursor = db.cursor()

    sql = ''' 
        SELECT
            X.id,
            X.date_event,
            L.id,
            L.name,
            E.id,
            S.description,
            B.description,
            E.description,
            X.minutes,
            X.note
        FROM
            EventXLine X
        INNER JOIN Event E ON
            X.id_event = E.id
        INNER JOIN User_table U ON
            X.id_user = U.id
        INNER JOIN Branch B ON
            E.id_branch = B.id
        INNER JOIN Sub_classification S ON
            B.id_sub_classification = S.id
        INNER JOIN Line L ON
            X.id_line = L.id
        WHERE
            X.date_event BETWEEN DATE_ADD(DATE(%s),
                INTERVAL 6 HOUR) AND DATE_ADD(DATE(%s),
                INTERVAL 6 HOUR)
    '''

    cursor.execute(sql, (begin, end))


    results = cursor.fetchall()


    cursor.close()

    db.close()

    return results


def beginDate(year, week):


    first = datetime.datetime(year, 1, 1)

    first = first - datetime.timedelta(days=first.weekday())
    
    begin = first + datetime.timedelta(days=7*(week-1))

    return begin

def endDate(begin):

    end = begin + datetime.timedelta(days=6)#dia comleto laboral con fin de semana

    return end


def printTable(results):
    for r in results:
        for c in r:
            print(str(c)+"\t", end="")
        print("\n")



def getTurn(date_event):
    if  6 <= date_event.hour and date_event.hour < 14:
        return 1
    elif 14 <= date_event.hour and date_event.hour < 22:
        return 2

    elif 22 <= date_event.hour and date_event.hour <= 23:
        return 3
    # este caso de turno 4 es cuando cambia de fecha y
    # es turno 3 del dia anterior
    else:
        return 4

def getCoord(date_event):
    turn = getTurn(date_event)
    day = date_event.weekday()

    if turn  == 4:
        turn = 3
        day -= 1

    return (day * 3) + turn  



def getFilterEventsBy(line, E):
    r = [] 

    for e in E:
        if e.line == line:
            r.append(e)
    
    return r



def engineTranslateCurrentData(dateBegin):
    Event = []
    Lines = []
    Filter = []
    Raw = []

    
    end = endDate(dateBegin)
    r = getEvents(dateBegin, end)
    l = getLines()

    # Agrega los paros de la BD de una semana a una lista de paros
    for e in r:
        Event.append(EventXLineHolder(e[0], e[1], e[2], e[3], e[4], e[5], e[6], e[7], float(e[8]), e[9]))


    # Obtener la lista de Lineas de la BD
    for e in l:
        Lines.append(Line(e[0], e[1]))
 
    print("\n\nProcessing  Lines...\n")

    # Mapper Algorithm
    for l in Lines:

        lineName = l.name

        eventMapper = EventMapper(lineName)


        # filtra los paros de una linea especifica y retorna una lista
        Filter = getFilterEventsBy(lineName, Event)
        

        # agregar los paros de forma mapeada de una linea especifica
        for e in Filter:

            code = e.id_event
            day = e.date_event.weekday()
            turn = getTurn(e.date_event)

            if turn == 4:
                if day != 0:
                    day -= 1  

                turn = 3 

            minute = e.minutes

            eventMapper.add(EventHolder(code, day, turn, float(minute)))

            last = len(eventMapper.Event) - 1

            eventMapper.Event[last].setTransaction(e.id_transaction)
            eventMapper.Event[last].print()

        Raw.append(eventMapper)


    return Raw




def engineCurrentData(dateBegin):

    Event = []
    Lines = []
    Filter = []
    Raw = []


    end = endDate(dateBegin)
    r = getEvents(dateBegin, end)
    l = getLines()

    # Agrega los paros de la BD de una semana a una lista de paros
    for e in r:
        Event.append(EventXLineHolder(e[0], e[1], e[2], e[3], e[4], e[5], e[6], e[7], float(e[8]), e[9]))


    # Obtener la lista de Lineas de la BD
    for e in l:
        Lines.append(Line(e[0], e[1]))
 
    print("\n\nProcessing  Lines...\n")

    # Mapper Algorithm
    for l in Lines:

        lineName = l.name

        rawMapper = RawMapper(lineName)


        # filtra los paros de una linea especifica y retorna una lista
        Filter = getFilterEventsBy(lineName, Event)
        
        lineMapper = LineMapper(lineName)


        # agregar los paros de forma mapeada de una linea especifica
        for e in Filter:
            lineMapper.add(e)


        rowEdge = 8
        colEdge = 2

        for i in range(len(lineMapper.sheet)):    
            rawMapper.codes.append( CodeHolder(rowEdge, 1, lineMapper.codes[i])) #fila, columna, minuto     
            for j in range(len(lineMapper.sheet[i])):

                coord = colEdge + getCoord(lineMapper.sheet[i][j].date_event)
                minutes = lineMapper.sheet[i][j].minutes

                rawMapper.add(DataHolder(rowEdge, coord, minutes))

            rowEdge += 1

        Raw.append(rawMapper)

    return Raw


    
def getStringMetaCoord(col):

    Week = ["L", "K", "M", "J", "V", "S", "D"]
    Turn = ["3", "2", "1"]

    edge = col - 2

    day = math.ceil(edge / 3) - 1 

    turn =  edge % 3

    return "" + Week[day] + "\t" + Turn[turn]




def printPrefab(Raw):

    for r in Raw:
        print("--%s--"%(r.line_name))

        print("lenCodes:%d\t|\tlenData:%d"%(len(r.codes), len(r.Data)))

        for data in r.Data:
            print("%s\t[%2d][%2d]->%d"%
            (getStringMetaCoord(data.col), data.row, data.col, data.minutes))



def initEngineOEE(file_template, dateBegin):


    Raw = engineCurrentData(dateBegin)


    # printPrefab(Raw)


    print("Escribiendo en plantilla...")

    #inicio secuancia de escritura a la plantilla
    workbook = load_workbook(filename=FILE_NAME)

    for r in Raw:

        sheet = workbook[r.line_name]

        for code in r.codes:
            sheet.cell(row=code.row, column=code.col).value = code.code
            
        for data in r.Data:
            sheet.cell(row=data.row, column=data.col).value = data.minutes
            

    workbook.save(filename=file_template)
    workbook.close()


    print("Listo!")
    return "Done"



def matchMapper(Before, After):

    for i in [12]:
        b = Before[i]
        a = After[0] 
    
        print("--MatchMapper--\n")

        print("lenCodes: %d == %d\t|\tlenData: %d == %d"%
        (len(b.codes), len(a.codes), len(b.Data), len(a.Data)))


def matchTest(file_template, dateBegin):

    # Before = engineCurrentData(dateBegin)
    # After = readTemplate(file_template, dateBegin)

    # matchMapper(Before, After)
    Lines = [] 
    l = getLines()

    # Obtener la lista de Lineas de la BD
    for e in l:
        Lines.append(Line(e[0], e[1]))
 


    Before = engineTranslateCurrentData(dateBegin)

    # for l in Before:
    #     l.print()


    # print("")

    After = readTemplate(file_template, dateBegin)

    # for l in After:
    #     l.print()

    # print("len>>\tB%d == A%d === L%d\n\n\n\n"%(len(Before), len(After), len(Lines)))

    for l in range(len(Lines)):
        print("\n\t-Line: %s-"%(Lines[l].name))
        i = 0 

        for e in After[l].Event:
            IN = sweepIn(Before[l].Event, e)
            
            OUT = Before[l].isIn(e)

            if OUT == -1:
                print("Out>>\t", end="")
                After[l].Event[i].print()
                insertEvent(After[l].Event[i], Lines[l].id, dateBegin)
            
            if IN != -1:
                print("In>>\t", end="")
                After[l].Event[IN].print()

                before = Before[l].Event[IN].minute
                after =  After[l].Event[IN].minute
                transaction = Before[l].Event[IN].transaction

                print("%d\t-->\t[%d]>>%d\n\n"
                %(before, transaction, after))

                updateEvent(After[l].Event[IN].minute, Before[l].Event[IN].transaction)

            i += 1

def insertEvent(data, lineID, dateBegin):
    
    db = pymysql.connect(IP, "brch", PASS, "mes")


    dayCalc = dateBegin + datetime.timedelta(days=data.day, hours=(data.turn * 8 - 1))
  

    print(dayCalc, "\n")

    cursor = db.cursor()
    sql = "INSERT INTO EventXLine(id_event, id_line, date_event, turn, minutes, id_user, note)"\
          " VALUES(%s, %s, %s, %s, %s, %s, %s)"

    cursor.execute(sql, (str(data.code), str(lineID), str(dayCalc), str(data.turn), float(data.minute), str(5), "Ajuste de tiempos"))

    db.commit()
    cursor.close()

    db.close()


def updateEvent(minute, transaction):
    db = pymysql.connect(IP, "brch", PASS, "mes")



    cursor = db.cursor()
    sql = "Update EventXLine set minutes=%s where id=%s"

    cursor.execute(sql, (float(minute), str(transaction)))

    db.commit()
    cursor.close()

    db.close()







def sweepIn(B, current):
    i = 0 

    for e in B:
        if e.code == current.code and e.day == current.day and e.turn == current.turn:
            diff  = e.minute - current.minute

            if diff != 0:
                return i
    

        i += 1


    return -1

def sweepOut(B, current):

    for e in B:
        if e.code != current.code and e.day != current.day and e.turn != current.turn:
          return True


    return False



def readTemplate(file_template, dateBegin):

    #abrir archivo
    workbook = load_workbook(filename=file_template)
    Raw = []
    Lines = []

    L = getLines()
    # Obtener la lista de Lineas de la BD
    for e in L:

        Lines.append(Line(e[0], e[1]))


    for l in  Lines:

        lineName = l.name

        sheet = workbook[lineName]

        # print(sheet.dimensions)
        # print("Minimum row: {0}".format(sheet.min_row))
        # print("Maximum row: {0}".format(sheet.max_row))
        # print("Minimum column: {0}".format(sheet.min_column))
        # print("Maximum column: {0}".format(sheet.max_column))

        rawMapper = RawMapper(lineName)
        eventMapper = EventMapper(lineName)

        rowEdge = 8
        colEdge = 2

        for row in sheet.iter_rows(min_row=rowEdge, min_col=1, max_col=23):
            pcol = 1
            for cell in row:
                code = 0 
                if cell.value:
                    if type(cell.value) is int:

                        if pcol == 1:
                            rawMapper.codes.append(CodeHolder(rowEdge, pcol, cell.value))
                        else:
                            code = rawMapper.codes[rawMapper.findCode(rowEdge)].code
                            day = math.ceil((pcol-2) / 3 ) - 1

                            calcTurn = (pcol-2) % 3 
                            turn = calcTurn if calcTurn != 0 else 3

                            minute = cell.value

                            eventMapper.add(EventHolder(code, day, turn, float(minute)))
                            # print("[%d]\t(%d, %d, %d, %d)"%(rowEdge, code, day, turn, minute))

                            rawMapper.add(DataHolder(rowEdge, pcol, cell.value))
                
                pcol += 1

            rowEdge += 1

        Raw.append(eventMapper)


    workbook.close()


    return Raw

if __name__ == "__main__":
    start = time.time()

    begin = beginDate(2020, 20)

    TEMP_FOLDER = "files/"

    file_name = "validacionOEE.xlsx"

    # initEngineOEE(TEMP_FOLDER + file_name, begin)

    matchTest(TEMP_FOLDER + file_name, begin)
   
    end = time.time()

    elapsed = (end - start) / 60
    print("Time elapsed: %f" % (elapsed))